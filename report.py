import os
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_report(jobs: list[dict], output_dir: str = "output") -> str:
    """Generate a formatted .xlsx report and return the file path."""
    Path(output_dir).mkdir(exist_ok=True)
    today = str(date.today())
    filepath = f"{output_dir}/data_jobs_{today}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "New Jobs"

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E79")
    cell_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="EBF3FB")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Data Roles — New Listings as of {today}"
    title_cell.font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Summary row ───────────────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Total new roles found: {len(jobs)}"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="595959")
    ws["A2"].alignment = left
    ws.row_dimensions[2].height = 18

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = ["#", "Company", "Job Title", "Department", "Location", "Date Found", "Apply Link"]
    col_widths = [5, 18, 40, 22, 25, 14, 50]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    # Sort by company then title
    jobs_sorted = sorted(jobs, key=lambda j: (j["company"].lower(), j["title"].lower()))

    for i, job in enumerate(jobs_sorted, start=1):
        row = i + 3
        fill = alt_fill if i % 2 == 0 else None

        values = [
            i,
            job.get("company", ""),
            job.get("title", ""),
            job.get("department", ""),
            job.get("location", ""),
            job.get("date_found", ""),
            job.get("url", ""),
        ]
        aligns = [center, left, left, left, left, center, left]

        for col, (val, align) in enumerate(zip(values, aligns), start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = cell_font
            cell.alignment = align
            cell.border = border
            if fill:
                cell.fill = fill

        # Hyperlink the Apply Link cell
        url = job.get("url", "")
        if url:
            link_cell = ws.cell(row=row, column=7)
            link_cell.hyperlink = url
            link_cell.value = "Apply →"
            link_cell.font = Font(name="Arial", size=10, color="1155CC", underline="single")

        ws.row_dimensions[row].height = 18

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:G{3 + len(jobs)}"

    wb.save(filepath)
    return filepath


if __name__ == "__main__":
    # Quick smoke test
    sample = [
        {
            "job_id": "test_1",
            "company": "Stripe",
            "title": "Senior Data Engineer",
            "department": "Data Platform",
            "location": "Remote, USA",
            "url": "https://stripe.com/jobs/test",
            "date_found": "2026-06-02",
        },
        {
            "job_id": "test_2",
            "company": "Airbnb",
            "title": "Analytics Engineer II",
            "department": "Data Science",
            "location": "San Francisco, CA",
            "url": "https://airbnb.com/careers/test",
            "date_found": "2026-06-02",
        },
    ]
    path = generate_report(sample)
    print(f"Report saved to: {path}")
